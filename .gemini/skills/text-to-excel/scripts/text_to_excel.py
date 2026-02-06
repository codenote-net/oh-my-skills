import argparse
import openpyxl
import os

def parse_markdown_table(text_content):
    """
    Parses a simple Markdown table from text content.
    Returns a list of lists, where each inner list is a row.
    """
    rows = []
    lines = text_content.strip().splitlines()
    for line in lines:
        line = line.strip()
        if not line.startswith('|') or not line.endswith('|'):
            continue
        
        # Check for and ignore separator line like |---| or |:--:|
        if all(c in '-:| ' for c in line[1:-1]):
            continue
            
        # Split by '|', remove the empty strings from the start and end, and strip each cell
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if cells:
            rows.append(cells)
    return rows

def write_to_excel(data, output_excel_path, template_excel_path=None):
    """
    Writes data (list of lists) to an Excel file.
    Optionally uses a template Excel file.
    """
    # Load workbook (template or new)
    if template_excel_path and os.path.exists(template_excel_path):
        workbook = openpyxl.load_workbook(template_excel_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    # Write data to the sheet
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_data in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_data)

    # Save the workbook
    workbook.save(output_excel_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert Markdown table text to an Excel file.")
    parser.add_argument("input_text_path", help="Path to the input text file containing a Markdown table.")
    parser.add_argument("output_excel_path", help="Path for the output Excel file (e.g., output.xlsx).")
    parser.add_argument("--template", dest="template_excel_path", help="Optional path to an Excel template file.", default=None)

    args = parser.parse_args()

    try:
        # Read input text
        with open(args.input_text_path, 'r', encoding='utf-8') as f:
            text_content = f.read()
        
        # Parse the markdown table
        parsed_data = parse_markdown_table(text_content)
        
        if not parsed_data:
            print("No Markdown table found in the input text.", file=os.stderr)
            exit(1)
            
        # Write the parsed data to Excel
        write_to_excel(parsed_data, args.output_excel_path, args.template_excel_path)
        
        print(f"Successfully processed '{args.input_text_path}' and saved to '{args.output_excel_path}'")

    except Exception as e:
        print(f"An error occurred: {e}", file=os.stderr)
        exit(1)